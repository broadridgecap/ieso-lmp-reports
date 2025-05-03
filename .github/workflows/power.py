import requests
import pandas as pd
from lxml import etree
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import defaultdict

#Load override date from Excel
output_dir = Path("output")
output_dir.mkdir(parents=True, exist_ok=True)
input_excel_path = output_dir / "ieso_lmp_input.xlsx"

# Create input file if missing
if not input_excel_path.exists():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Override Date"
    ws["B1"] = ""  # Leave blank to use today's date
    wb.save(input_excel_path)

#date override
wb_input = load_workbook(input_excel_path)
ws_input = wb_input.active
override_date = ws_input["B1"].value

if override_date:
    override_date = str(override_date).strip()
    if not override_date.isdigit() or len(override_date) != 8:
        print("⚠️ Invalid override date format in B1. Expected YYYYMMDD. Using today's date instead.")
        override_date = None
    else:
        print(f"📅 Using override date: {override_date}")
else:
    print("ℹ️ No override date in B1. Using today's date.")

#correct IESO XML URL 
if override_date:
    date_for_url = override_date
    url = f"https://reports-public.ieso.ca/public/PredispHourlyIntertieLMP/PUB_PredispHourlyIntertieLMP_{date_for_url}.xml"
else:
    date_for_url = datetime.today().strftime("%Y%m%d")
    url = "https://reports-public.ieso.ca/public/PredispHourlyIntertieLMP/PUB_PredispHourlyIntertieLMP.xml"

print(f"📡 Fetching data from: {url}")

# === Step 3: Download XML File ===
try:
    response = requests.get(url)
    response.raise_for_status()
    print("✅ XML file downloaded.")
except Exception as e:
    print("❌ Failed to download XML:", e)
    exit(1)

#parse XML and Extract CreatedAt
ns = {"ieso": "http://www.ieso.ca/schema"}
tree = etree.parse(BytesIO(response.content))
root = tree.getroot()
created_at = root.findtext(".//ieso:CreatedAt", namespaces=ns)

#records from XML 
records = []
for intertie in root.findall(".//ieso:IntertieLMPrice", namespaces=ns):
    intertie_name = intertie.find("ieso:IntertiePLName", namespaces=ns).text
    for component in intertie.findall("ieso:Components", namespaces=ns):
        component_name = component.find("ieso:LMPComponent", namespaces=ns).text
        for hour_entry in component.findall("ieso:HourlyLMP", namespaces=ns):
            hour = int(hour_entry.find("ieso:Hour", namespaces=ns).text)
            lmp = float(hour_entry.find("ieso:LMP", namespaces=ns).text)
            records.append({
                "Intertie Pricing Location": intertie_name,
                "Intertie LMP & Components": component_name,
                "Hour": hour,
                "LMP": lmp
            })

#format data
df = pd.DataFrame(records)
component_order = [
    "Intertie LMP",
    "Energy Congestion Price",
    "Energy Loss Price",
    "External Congestion Price",
    "Net Interchange Scheduling Limit (NISL) Price"
]
df["Intertie LMP & Components"] = pd.Categorical(
    df["Intertie LMP & Components"],
    categories=component_order,
    ordered=True
)
df.sort_values(by=["Intertie Pricing Location", "Intertie LMP & Components"], inplace=True)

#pivot data
pivot_df = df.pivot_table(
    index=["Intertie Pricing Location", "Intertie LMP & Components"],
    columns="Hour", values="LMP", observed=False
).reset_index()

#timestamp
timestamp = datetime.now().strftime("%H%M%S")
safe_output_path = output_dir / f"ieso_lmp_final_{date_for_url}_{timestamp}.xlsx"
pivot_df.to_excel(safe_output_path, index=False)

#format excel
wb = load_workbook(safe_output_path)
ws = wb.active

ws.insert_rows(1)
ws.insert_rows(1)

ws["A1"] = f"CreatedAt: {created_at}"
ws["A1"].font = Font(bold=True)
ws.merge_cells("A1:Z1")

start_col, end_col = 3, 26
start_letter = get_column_letter(start_col)
end_letter = get_column_letter(end_col)
ws.merge_cells(f"{start_letter}2:{end_letter}2")
ws[f"{start_letter}2"] = "Predispatch Hourly Energy LMP ($/MWh) for hour"
ws[f"{start_letter}2"].font = Font(bold=True)

for cell in ws[3]:
    cell.font = Font(bold=True)

# row styling by Intertie
color_palette = [
    "FFDDDD", "DDEEFF", "CCFFCC", "FFFFCC", "FFCCE5",
    "E0E0E0", "FFCC99", "CCE5FF", "D1C4E9", "FFF3E0"
]
intertie_colors = defaultdict(lambda: PatternFill(fill_type="solid", fgColor="FFFFFF"))
unique_interties = list(pivot_df["Intertie Pricing Location"].unique())
for i, name in enumerate(unique_interties):
    intertie_colors[name] = PatternFill(fill_type="solid", fgColor=color_palette[i % len(color_palette)])


thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for row in range(4, ws.max_row + 1):
    intertie_value = ws.cell(row=row, column=1).value
    fill = intertie_colors[intertie_value]
    for col in range(1, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == 1:
            cell.fill = fill
        cell.border = thin_border

for row in [1, 2, 3]:
    for col in range(1, end_col + 1):
        ws.cell(row=row, column=col).border = thin_border

wb.save(safe_output_path)
print(f"✅ Final Excel report saved to: {safe_output_path}")
